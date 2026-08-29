#!/usr/bin/env python3
"""build-xlsx.py - accumulate scan results into one growing Excel workbook.

One file (default data/matches.xlsx) with one sheet per day, named by the day
(Excel forbids "/" in tab names, so a date is written as e.g. 29-8-2026).
Inside each day's sheet, every scan run of that day is appended as a set of
stacked tables (one below the other), each labelled with its run time, so a
full day with 6 runs shows 6 stacked match tables (plus their fuller lists).

If today's sheet already exists (a later run of the same day), we append a new
stacked table below the previous ones instead of overwriting. Tomorrow creates
a fresh sheet in the same file, so the workbook grows across days.

Inputs (env):
  MATCHES_CSV       enriched per-run matches (from generate-alert.mjs)
  CANDIDATES_JSON   relevant candidates this run (from scan.mjs)
  OUT_XLSX          path to the accumulating workbook (default data/matches.xlsx)
  RUN_LABEL         label for this run's table (e.g. "13:00"); optional
  TZ_OFFSET_H       hours offset for local-day naming (default 2, Tripoli)
"""

import csv
import json
import os
import sys
from datetime import datetime, timedelta, timezone
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

matches_csv = os.environ.get("MATCHES_CSV", "/tmp/alert-rows.csv")
candidates_json = os.environ.get("CANDIDATES_JSON", "data/candidates.json")
all_found_json = os.environ.get("ALL_FOUND_JSON", "data/all-found.json")
scan_runs_tsv = os.environ.get("SCAN_RUNS_TSV", "data/scan-runs.tsv")
out_xlsx = os.environ.get("OUT_XLSX", "data/matches.xlsx")
run_label = os.environ.get("RUN_LABEL", "").strip()
tz_offset = float(os.environ.get("TZ_OFFSET_H", "2"))

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
RUN_FILL = PatternFill("solid", fgColor="DDEBF7")
RUN_FONT = Font(bold=True, size=12, color="1F4E79")
LINK_FONT = Font(color="0563C1", underline="single")
URGENT_FILL = PatternFill("solid", fgColor="FFEB9C")
THIN = Side(style="thin", color="D0D0D0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center")


def local_now():
    return datetime.now(timezone.utc) + timedelta(hours=tz_offset)


def day_name(dt):
    # Excel tab names cannot contain "/" -> use dashes: 29-8-2026
    return f"{dt.day}-{dt.month}-{dt.year}"


def style_header(ws, row, headers):
    for c, _ in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=c, value=headers[c - 1])
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT


def write_row(ws, row, values, headers):
    for c in range(1, len(headers) + 1):
        ws.cell(row=row, column=c).value = values[c - 1]
        ws.cell(row=row, column=c).border = BORDER


def write_link(ws, row, col, url, display=None):
    cell = ws.cell(row=row, column=col, value=display or url)
    if url and url.startswith("http"):
        cell.hyperlink = url
        cell.font = LINK_FONT
    return cell


def read_matches():
    if not os.path.exists(matches_csv):
        return []
    with open(matches_csv, newline="", encoding="utf-8-sig") as f:
        return list(csv.DictReader(f))


def read_candidates():
    if not os.path.exists(candidates_json):
        return []
    try:
        with open(candidates_json, encoding="utf-8-sig") as f:
            return json.load(f).get("candidates", [])
    except Exception:
        return []


def read_all_found():
    """Read the wide 'all found jobs' (700+) export. Returns (list, count)."""
    if not os.path.exists(all_found_json):
        return [], 0
    try:
        with open(all_found_json, encoding="utf-8-sig") as f:
            d = json.load(f)
            return d.get("offers", []), d.get("count", 0)
    except Exception:
        return [], 0


STAT_HDRS = ["Local time", "Sites pulled", "Search results", "Removed by title",
             "Removed by location", "Duplicates", "New matches added"]

ALL_HDRS = ["Company", "Role", "Location", "Posted", "Pay", "Source", "Apply link"]


def write_all_found_sheet(ws, all_offers, now, label):
    """Write (or refresh) the wide '700+' sheet listing every raw job found this
    run across all sources. Unlike the strict CV-fit sheet, this is the full
    browse deluge. We refresh a single 'All Jobs 700+' sheet each run so it shows
    the latest full pull; the strict per-day sheets already accumulate history.
    """
    # Clear previous content for this sheet so it reflects this run's full pull.
    ws.delete_rows(1, ws.max_row or 1)
    ws.cell(row=1, column=1, value=(
        f"Every job found this run across all sources ({len(all_offers)} total), "
        "before strict CV filtering. Browse these to search wider."))
    ws.cell(row=1, column=1).font = Font(italic=True, size=9, color="808080")
    style_header(ws, 2, ALL_HDRS)
    row = 3
    for c in all_offers:
        posted = ""
        if c.get("postedAt"):
            posted = c["postedAt"][:10] if isinstance(c["postedAt"], str) else str(c.get("postedAt"))
        pay = ""
        sal = c.get("salary")
        if isinstance(sal, dict):
            lo, hi = sal.get("min"), sal.get("max")
            cur = sal.get("currency", "")
            if lo and hi:
                pay = f"{lo}-{hi} {cur}".strip()
            elif lo:
                pay = f"{lo} {cur}".strip()
        write_row(ws, row, [
            c.get("company", ""), c.get("role", ""), c.get("location", ""),
            posted, pay, c.get("source", ""), c.get("url", ""),
        ], ALL_HDRS)
        write_link(ws, row, 7, c.get("url", ""))
        row += 1
        if row > 1048570:
            break
    return row



def append_scan_stats(ws, now):
    """Append this run's scan-statistics row to the shared 'Scan Stats' sheet.

    Reads the latest completed row of data/scan-runs.tsv (companies+boards =
    sites pulled, found = total search results, filtered_title / filtered_location
    / dupes = funnel, new_added = survivors). This is the 'how many searches and
    how many sites' view the user asked for. Returns nothing; ignores missing file.
    """
    if not os.path.exists(scan_runs_tsv):
        return
    import csv as _csv
    with open(scan_runs_tsv, newline="", encoding="utf-8-sig") as f:
        rows = list(_csv.DictReader(f, delimiter="\t"))
    if not rows:
        return

    def _i(d, k):
        try:
            return int(float(d.get(k) or 0))
        except (TypeError, ValueError):
            return 0

    # Take the last completed run of this scan as today's current stats row.
    last = rows[-1]
    sites = _i(last, "companies") + _i(last, "boards")
    stats = [
        now.strftime("%Y-%m-%d %H:%M"),
        sites,
        _i(last, "found"),
        _i(last, "filtered_title") + _i(last, "filtered_content") +
        _i(last, "filtered_tier") + _i(last, "filtered_posting_age") +
        _i(last, "filtered_salary") + _i(last, "filtered_blacklist") +
        _i(last, "filtered_visa") + _i(last, "filtered_country_eligibility"),
        _i(last, "filtered_location"),
        _i(last, "dupes") + _i(last, "filtered_cooldown"),
        _i(last, "new_added"),
    ]

    if "Scan Stats" not in ws.parent.sheetnames:
        s = ws.parent.create_sheet("Scan Stats")
        s.cell(row=1, column=1, value=("Per-run scan statistics: how many job "
                                       "searches were pulled, from how many sites, and what survived filtering."))
        s.cell(row=1, column=1).font = Font(italic=True, size=9, color="808080")
        style_header(s, 2, STAT_HDRS)
        s._next_row = 3
    else:
        s = ws.parent["Scan Stats"]
        if not hasattr(s, "_next_row"):
            # Recompute the next empty row if this sheet was freshly loaded.
            nxt = 1
            while s.cell(row=nxt, column=1).value not in (None, ""):
                nxt += 1
            s._next_row = nxt

    r = s._next_row
    write_row(s, r, stats, STAT_HDRS)
    s._next_row = r + 1


MATCH_HDRS = ["Fit", "Field", "Company", "Role", "Location", "Pay", "Age", "Urgency", "Apply link", "Source"]
FULL_HDRS = ["Company", "Role", "Location", "Posted", "Pay", "Source", "Apply link"]

# Fit-score floor for the strict apply-worthy sheet. On-site blocked-hub roles
# are hard-capped at <=15 by generate-alert, so this floor drops them AND the
# low near-miss noise, while leaving a few easy-to-eyeball 40+ rows. Balanced,
# per the user's "few easy near-misses" preference.
APP_MIN_FIT = 40


def apply_worthy(m):
    """Return True when a row is genuinely worth Waleed's click on the strict
    apply-worthy sheet: it is eligible (not an on-site blocked-hub role) and its
    fit score clears the balanced floor. Missing eligibility is treated as
    eligible (older rows / defensible default); missing score is kept so we
    never silently drop an undated or un-scored but otherwise strong role."""
    elig = str(m.get("eligible", "yes")).strip().lower()
    if elig == "no":
        return False
    try:
        score = float(m.get("fit_score"))
    except (TypeError, ValueError):
        return True
    return score >= APP_MIN_FIT



def append_match_table(ws, row, matches, label):
    """Write a run's match table starting at `row`. Returns next free row.

    This is the strict APPLY-WORTHY sheet, so we show only rows that are genuinely
    worth a click and drop the noise that would waste Waleed's time:
      - skip on-site blocked-hub roles outright (eligible == "no" — these were
        hard-capped by generate-alert and are not eligible for him)
      - skip fit scores below APP_MIN_FIT (low / near-miss noise that is not a
        real match). Balanced by design: keeping rows from ~40 up leaves the
        few easy-to-eyeball near-misses without burying the real matches.
    The wider "Full List" table below still carries every relevant role for
    browsing, so nothing is hidden, only de-prioritised here.
    """
    if not matches:
        return row
    kept = [m for m in matches if apply_worthy(m)]
    if not kept:
        return row
    cell = ws.cell(row=row, column=1,
                   value=f"{label}  -  Apply-worthy matches")
    cell.font = RUN_FONT
    cell.fill = RUN_FILL
    row += 1
    ws.cell(row=row, column=1, value=("The jobs this run found that are genuinely worth your time: strong "
                                      "fits that are remote or in Libya/the Middle East, filtered so you don't waste clicks."))
    ws.cell(row=row, column=1).font = Font(italic=True, size=9, color="808080")
    row += 1
    style_header(ws, row, MATCH_HDRS)
    row += 1
    for m in kept:
        write_row(ws, row, [
            m.get("fit_score", ""), m.get("field", ""), m.get("company", ""),
            m.get("title", ""), m.get("location", ""), m.get("pay", ""),
            m.get("age", ""), m.get("urgency", ""), m.get("url", ""),
            m.get("source_site", ""),
        ], MATCH_HDRS)
        if str(m.get("urgency", "")).lower() == "urgent":
            for c in range(1, len(MATCH_HDRS) + 1):
                ws.cell(row=row, column=c).fill = URGENT_FILL
        write_link(ws, row, 9, m.get("url", ""))
        ws.cell(row=row, column=1).alignment = CENTER
        ws.cell(row=row, column=8).alignment = CENTER
        row += 1
    return row


def append_full_table(ws, row, candidates, label):
    """Write a run's fuller candidate list starting at `row`. Returns next free row."""
    if not candidates:
        return row
    cell = ws.cell(row=row, column=1,
                   value=f"{label}  -  Full List (all relevant jobs this run)")
    cell.font = RUN_FONT
    cell.fill = RUN_FILL
    row += 1
    ws.cell(row=row, column=1, value=("Every posting that met the relevance "
                                      "filters this run, so you can browse wider than the matches above."))
    ws.cell(row=row, column=1).font = Font(italic=True, size=9, color="808080")
    row += 1
    style_header(ws, row, FULL_HDRS)
    row += 1
    for c in candidates:
        posted = ""
        if c.get("postedAt"):
            posted = c["postedAt"][:10] if isinstance(c["postedAt"], str) else str(c.get("postedAt"))
        pay = ""
        sal = c.get("salary")
        if isinstance(sal, dict):
            lo, hi = sal.get("min"), sal.get("max")
            cur = sal.get("currency", "")
            if lo and hi:
                pay = f"{lo}-{hi} {cur}".strip()
            elif lo:
                pay = f"{lo} {cur}".strip()
        write_row(ws, row, [
            c.get("company", ""), c.get("role", ""), c.get("location", ""),
            posted, pay, c.get("source", ""), c.get("url", ""),
        ], FULL_HDRS)
        write_link(ws, row, 7, c.get("url", ""))
        row += 1
    return row


def main():
    now = local_now()
    label = run_label or now.strftime("%H:%M")
    name = day_name(now)
    matches = read_matches()
    candidates = read_candidates()
    all_found, all_count = read_all_found()

    if os.path.exists(out_xlsx):
        wb = load_workbook(out_xlsx)
    else:
        wb = Workbook()

    # Wide "All Jobs 700+" browse sheet: build it only ONCE per day (the first
    # run of the day). Later runs keep the same full 700+ pull so they don't
    # waste time re-listing ~2,800 rows. The day it was built is stored in a
    # hidden marker cell (H1); when the marker matches today we skip the rebuild.
    if "All Jobs 700+" in wb.sheetnames:
        ws_all = wb["All Jobs 700+"]
        marker = ws_all.cell(row=1, column=8).value
        if not marker or str(marker) != name:
            write_all_found_sheet(ws_all, all_found, now, label)
            ws_all.cell(row=1, column=8, value=name)
    else:
        # Reuse the default empty "Sheet" as the wide sheet on a fresh file so
        # no leftover blank tab remains.
        dflt = "Sheet" if (len(wb.sheetnames) == 1 and wb.sheetnames[0] == "Sheet"
                           and (wb["Sheet"].max_row or 0) <= 1) else None
        ws_all = wb[dflt] if dflt else wb.create_sheet("All Jobs 700+")
        if dflt:
            ws_all.title = "All Jobs 700+"
        write_all_found_sheet(ws_all, all_found, now, label)
        ws_all.cell(row=1, column=8, value=name)

    # Get (or create) today's sheet.
    if name in wb.sheetnames:
        ws = wb[name]
    else:
        # Reuse the default empty "Sheet" as today's sheet so fresh files
        # don't carry a leftover blank tab; otherwise create a new one.
        if len(wb.sheetnames) == 1 and wb.sheetnames[0] == "Sheet" and (wb["Sheet"].max_row or 0) <= 1:
            ws = wb["Sheet"]
            ws.title = name
        else:
            ws = wb.create_sheet(name)

    # Determine the first empty row (after any existing content).
    # A blank sheet reports max_row=1; we then start tables at row 1.
    row = 1 if ws.max_row <= 1 else ws.max_row + 1

    # Write this run's two stacked tables.
    row = append_match_table(ws, row, matches, label)
    row = append_full_table(ws, row, candidates, label)

    # Leave a blank spacer row between runs so tables stay visually separate.
    ws.cell(row=row, column=1, value="")

    # Record this run's scan statistics (searches pulled, sites, funnel) once.
    append_scan_stats(ws, now)

    out = Path(out_xlsx)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    print(f"xlsx updated: {out}  sheet={name} run={label} matches={len(matches)} candidates={len(candidates)} all_found={len(all_found)}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
